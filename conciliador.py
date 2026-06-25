# conciliador.py
# Este archivo contiene toda la lógica de negocio:
# Lectura, limpieza, creación de ID, comparación, generación del reporte Excel con formato profesional.


# Módulo para conciliación de dos archivos PLE Compras (comparación de series y números)
# Autor: Yessly Poma
# Fecha: 2026-06-16


import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time


# ------------------------------------------------------------
# 0. NORMALIZACIÓN DE TIPO DE DOCUMENTO (COLUMNA G)
# ------------------------------------------------------------


def normalizar_tipo_documento(valor):
    """
    Normaliza el tipo de documento de la columna G (índice 6).
    - Convierte a string
    - Elimina decimales (42.0 -> 42)
    - Rellena con cero a la izquierda para tener 2 dígitos
    - Maneja casos especiales (NaN, vacío, etc.)
   
    Ejemplos:
        1   -> '01'
        03  -> '03'
        42.0 -> '42'
        0.0 -> '00'
        NaN -> '00'
        ''  -> '00'
    """
    if pd.isna(valor) or str(valor).strip() == '':
        return '00'
   
    valor_str = str(valor).strip()
   
    if '.' in valor_str:
        try:
            valor_str = str(int(float(valor_str)))
        except:
            pass
   
    import re
    digitos = re.search(r'(\d+)', valor_str)
    if digitos:
        valor_str = digitos.group(1)
    else:
        return '00'
   
    return valor_str.zfill(2)




# ------------------------------------------------------------
# 0.1. FUNCIÓN PARA OBTENER MES DE UNA FECHA
# ------------------------------------------------------------


def extraer_mes_fecha(fecha_str):
    """
    Extrae el mes y año de una fecha en formato DD/MM/YYYY o similar.
    Retorna un string en formato 'YYYYMM' para comparación.
    """
    if pd.isna(fecha_str) or str(fecha_str).strip() == '':
        return None
   
    fecha_str = str(fecha_str).strip()
   
    # Intentar diferentes formatos de fecha
    formatos = [
        '%d/%m/%Y',
        '%d/%m/%y',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%m/%d/%Y'
    ]
   
    for formato in formatos:
        try:
            fecha = pd.to_datetime(fecha_str, format=formato)
            return fecha.strftime('%Y%m')
        except:
            continue
   
    # Si falla, intentar con el parser flexible de pandas
    try:
        fecha = pd.to_datetime(fecha_str)
        return fecha.strftime('%Y%m')
    except:
        return None




# ------------------------------------------------------------
# 1. LECTURA Y NORMALIZACIÓN DE UN ARCHIVO
# ------------------------------------------------------------


def leer_archivo_conciliacion(ruta_archivo, hoja=None, fila_inicio=1):
    skip_rows = fila_inicio - 1


    if hoja is None:
        df = pd.read_excel(ruta_archivo, header=None, skiprows=skip_rows)
    else:
        df = pd.read_excel(ruta_archivo, sheet_name=hoja, header=None, skiprows=skip_rows)


    if df.empty:
        raise ValueError("No se encontraron datos en el archivo u hoja especificada.")


    if df.shape[1] < 13:  # Necesitamos hasta columna M (índice 12)
        raise ValueError("El archivo no tiene suficientes columnas (se necesitan G, H, J y M).")


    df_proc = df.copy()
   
    # --- Extraer y normalizar Tipo de Documento (columna G, índice 6) ---
    tipo_raw = df_proc.iloc[:, 6] if df_proc.shape[1] > 6 else pd.Series(['00'] * len(df_proc))
    df_proc['TipoDoc_Norm'] = tipo_raw.apply(normalizar_tipo_documento)
   
    # --- Extraer Serie (columna H, índice 7) ---
    serie_raw = df_proc.iloc[:, 7].astype(str).str.strip()
   
    # --- Extraer y normalizar Número (columna J, índice 9) ---
    numero_raw = df_proc.iloc[:, 9].astype(str).str.strip()
    numero_clean = numero_raw.str.replace(r'\.0$', '', regex=True)
    numero_fill = numero_clean.str.zfill(8)
   
    # --- Extraer RUC (columna M, índice 12) ---
    ruc_raw = df_proc.iloc[:, 12].astype(str).str.strip() if df_proc.shape[1] > 12 else pd.Series([''] * len(df_proc))


    df_proc['Tipo'] = tipo_raw
    df_proc['Serie'] = serie_raw
    df_proc['Numero'] = numero_fill
    df_proc['RUC'] = ruc_raw
   
    # ID de conciliación más completo: Tipo + Serie + Numero + RUC
    df_proc['ID_CONCILIACION'] = df_proc['TipoDoc_Norm'] + df_proc['Serie'] + df_proc['Numero'] + df_proc['RUC']
    df_proc['Hoja_Origen'] = hoja if hoja else 'Hoja1'


    return df_proc


def leer_todas_hojas_conciliacion(ruta_archivo, fila_inicio=1):
    """
    Lee todas las hojas del archivo, aplicando el mismo fila_inicio a cada una.
    """
    xlsx = pd.ExcelFile(ruta_archivo)
    hojas = xlsx.sheet_names
    dfs = []
    for hoja in hojas:
        try:
            df_hoja = leer_archivo_conciliacion(ruta_archivo, hoja=hoja, fila_inicio=fila_inicio)
            dfs.append(df_hoja)
        except Exception as e:
            print(f"Error en hoja {hoja}: {e}")
            continue
    if not dfs:
        raise ValueError("No se pudo leer ninguna hoja válida.")
    return pd.concat(dfs, ignore_index=True)


# ------------------------------------------------------------
# 2. COMPARACIÓN DE DOS ARCHIVOS
# ------------------------------------------------------------


def conciliar_archivos(df1, df2, nombre1="SIRE_SUNAT", nombre2="SIRE_BN"):
    """
    Compara dos DataFrames (ya procesados con ID_CONCILIACION) y devuelve:
        - resumen: dict con estadísticas
        - solo1: DataFrame con filas de df1 cuyo ID no está en df2
        - solo2: DataFrame con filas de df2 cuyo ID no está en df1
    """
    ids1 = set(df1['ID_CONCILIACION'].dropna().unique())
    ids2 = set(df2['ID_CONCILIACION'].dropna().unique())


    comunes = ids1 & ids2
    solo1_ids = ids1 - ids2
    solo2_ids = ids2 - ids1


    solo1 = df1[df1['ID_CONCILIACION'].isin(solo1_ids)].copy()
    solo2 = df2[df2['ID_CONCILIACION'].isin(solo2_ids)].copy()


    resumen = {
        'nombre_archivo_1': nombre1,
        'nombre_archivo_2': nombre2,
        'total_registros_1': len(df1),
        'total_registros_2': len(df2),
        'ids_unicos_1': len(ids1),
        'ids_unicos_2': len(ids2),
        'solo_en_1': len(solo1),
        'solo_en_2': len(solo2),
        'comunes': len(comunes),
        'diferencias_totales': len(solo1) + len(solo2)
    }


    return resumen, solo1, solo2


# ------------------------------------------------------------
# 3. CONTAR POR TIPO DE DOCUMENTO
# ------------------------------------------------------------


def contar_por_tipo_documento(df1, df2):
    """
    Cuenta los registros por tipo de documento usando la columna normalizada 'TipoDoc_Norm'.
    Los encabezados siempre serán 'Total SIRE_SUNAT' y 'Total SIRE_BN'.
    """
    TIPOS_DOC = {
        '00': '00', '01': '01', '02': '02', '03': '03', '04': '04',
        '05': '05', '06': '06', '07': '07', '08': '08', '09': '09',
        '12': '12', '14': '14', '16': '16', '30': '30', '42': '42',
        '46': '46', '53': '53'
    }
   
    if 'TipoDoc_Norm' in df1.columns:
        tipos1 = df1['TipoDoc_Norm']
    else:
        tipos1 = df1.iloc[:, 6].apply(normalizar_tipo_documento) if df1.shape[1] > 6 else pd.Series(['00'] * len(df1))
   
    if 'TipoDoc_Norm' in df2.columns:
        tipos2 = df2['TipoDoc_Norm']
    else:
        tipos2 = df2.iloc[:, 6].apply(normalizar_tipo_documento) if df2.shape[1] > 6 else pd.Series(['00'] * len(df2))
   
    conteo1 = tipos1.value_counts().to_dict()
    conteo2 = tipos2.value_counts().to_dict()
   
    todos_tipos = set(conteo1.keys()) | set(conteo2.keys())
   
    def orden_tipo(t):
        try:
            return int(t)
        except:
            return 999
   
    resultados = []
    for tipo in sorted(todos_tipos, key=orden_tipo):
        total1 = conteo1.get(tipo, 0)
        total2 = conteo2.get(tipo, 0)
        diferencia = abs(total1 - total2)
       
        if total1 > total2:
            libro_electronico = 'SIRE_SUNAT'
        elif total2 > total1:
            libro_electronico = 'SIRE_BN'
        else:
            libro_electronico = 'IGUALES'
       
        resultados.append({
            'Tipo de Documento': tipo,
            'Total SIRE_SUNAT': total1,
            'Total SIRE_BN': total2,
            'Diferencia': diferencia,
            'Libro Electrónico': libro_electronico
        })
   
    total1 = sum(r['Total SIRE_SUNAT'] for r in resultados)
    total2 = sum(r['Total SIRE_BN'] for r in resultados)
    total_fila = {
        'Tipo de Documento': 'TOTAL GENERAL',
        'Total SIRE_SUNAT': total1,
        'Total SIRE_BN': total2,
        'Diferencia': abs(total1 - total2),
        'Libro Electrónico': 'SIRE_SUNAT' if total1 > total2 else ('SIRE_BN' if total2 > total1 else 'IGUALES')
    }
    resultados.append(total_fila)
   
    return pd.DataFrame(resultados)


# ------------------------------------------------------------
# 4. GENERACIÓN DE REPORTE EXCEL
# ------------------------------------------------------------


def generar_reporte_conciliacion(resumen, solo1, solo2, df1, df2, nombre1, nombre2, ruta_salida):
    """
    Crea un archivo Excel con tres hojas:
        1. Resumen (dos tablas: resumen general + desglose por tipo)
        2. Solo en SIRE_SUNAT
        3. Solo en SIRE_BN
        4. En ambos SIRE_SUNAT y SIRE_BN

    Los encabezados siempre usan 'SIRE_SUNAT' y 'SIRE_BN' en lugar de los nombres de archivo.
    """
    wb = Workbook()
    wb.remove(wb.active)
   
    # ---------- Hoja 1: Resumen ----------
    ws1 = wb.create_sheet("Resumen")
   
    # ===== TABLA 1: Resumen General =====
    ws1['A1'] = 'RESUMEN GENERAL DE CONCILIACIÓN'
    ws1.merge_cells('A1:C1')
    ws1['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws1['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
   
    # ENCABEZADOS FIJOS: SIRE y PLE (no los nombres de archivo)
    headers_resumen = ['Concepto', 'SIRE_SUNAT', 'SIRE_BN']
    ws1.append(headers_resumen)
   
    datos_resumen = [
        ['Total registros (serie)', resumen['total_registros_1'], resumen['total_registros_2']],
        ['IDs únicos', resumen['ids_unicos_1'], resumen['ids_unicos_2']],
        ['Registros solo en este archivo', resumen['solo_en_1'], resumen['solo_en_2']],
        ['Registros en común', resumen['comunes'], resumen['comunes']],
        ['Diferencias totales', resumen['diferencias_totales'], resumen['diferencias_totales']]
    ]
    for fila in datos_resumen:
        ws1.append(fila)
   
    _aplicar_formato_encabezado(ws1, row=2)
   
    for col in ['A', 'B', 'C']:
        ws1.column_dimensions[col].width = 30
   
    # ===== Espacio entre tablas =====
    fila_actual = len(datos_resumen) + 4
   
    # ===== TABLA 2: Desglose por Tipo de Documento =====
    ws1.cell(row=fila_actual, column=1, value='DESGLOSE POR TIPO DE DOCUMENTO')
    ws1.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=5)
    ws1.cell(row=fila_actual, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws1.cell(row=fila_actual, column=1).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws1.cell(row=fila_actual, column=1).alignment = Alignment(horizontal='center', vertical='center')
    fila_actual += 1
   
    # Obtener desglose por tipo con encabezados fijos SIRE/PLE
    df_tipos = contar_por_tipo_documento(df1, df2)
   
    headers_tipos = list(df_tipos.columns)
    for col_idx, header in enumerate(headers_tipos, 1):
        ws1.cell(row=fila_actual, column=col_idx, value=header)
    fila_actual += 1
   
    for _, row in df_tipos.iterrows():
        for col_idx, header in enumerate(headers_tipos, 1):
            ws1.cell(row=fila_actual, column=col_idx, value=row[header])
        fila_actual += 1
   
    _aplicar_formato_encabezado(ws1, row=fila_actual - len(df_tipos) - 1)
   
    for col_idx in range(1, len(headers_tipos) + 1):
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = 25
   
    for row in range(fila_actual - len(df_tipos), fila_actual):
        for col_idx in [2, 3, 4]:
            if col_idx <= len(headers_tipos):
                cell = ws1.cell(row=row, column=col_idx)
                cell.number_format = '#,##0'
   
    # ---------- Hoja 2: Solo en SIRE_SUNAT ----------
    if not solo1.empty:
        ws2 = wb.create_sheet("Solo en SIRE_SUNAT")
        columnas_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                         [c for c in solo1.columns if c not in ['ID_CONCILIACION',  'Hoja_Origen']]
        df_export = solo1[columnas_orden]
        _escribir_dataframe_con_formato(ws2, df_export)
    else:
        ws2 = wb.create_sheet("Solo en SIRE_SUNAT")
        ws2.append(["No hay registros solo en SIRE_SUNAT"])


    # ---------- Hoja 3: Solo en SIRE_BN ----------
    if not solo2.empty:
        ws3 = wb.create_sheet("Solo en SIRE_BN")
        columnas_orden = ['ID_CONCILIACION',  'Hoja_Origen'] + \
                         [c for c in solo2.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
        df_export = solo2[columnas_orden]
        _escribir_dataframe_con_formato(ws3, df_export)
    else:
        ws3 = wb.create_sheet("Solo en SIRE_BN")
        ws3.append(["No hay registros solo en SIRE_BN"])


    wb.save(ruta_salida)
    return ruta_salida


# ------------------------------------------------------------
# 5. FUNCIONES AUXILIARES DE FORMATO
# ------------------------------------------------------------


def _aplicar_formato_encabezado(ws, row=1):
    """Aplica formato de encabezado (negrita, azul, centrado) a una fila específica."""
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for cell in ws[row]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border


def _escribir_dataframe_con_formato(ws, df, start_row=1):
    """
    Escribe un DataFrame en una hoja de cálculo a partir de la fila start_row,
    aplicando formato a los encabezados (que estarán en start_row).
    """
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)


    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


    _aplicar_formato_encabezado(ws, row=start_row)


    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_len + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def _escribir_dataframe_con_formato_con_titulo(ws, df, start_row, titulo, total=None):
    """
    Escribe un DataFrame en una hoja de cálculo a partir de la fila start_row,
    con un título antes de la tabla.
    """
    # Escribir título
    if titulo:
        ws.merge_cells(f'A{start_row}:E{start_row}')
        celda_titulo = ws.cell(row=start_row, column=1, value=titulo)
        celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        celda_titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1
   
    # Escribir encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
   
    # Escribir datos
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
   
    # Aplicar formato a los encabezados
    _aplicar_formato_encabezado(ws, row=start_row)
   
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_len + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
   
    # Retornar la siguiente fila disponible
    return start_row + len(df) + 2  # +2 por el espacio después de la tabla


# ------------------------------------------------------------
# 6. REPORTE: REGISTROS DE SIRE_BN (DESDE SIRE_BN HACIA SIRE_SUNAT)
# ------------------------------------------------------------
def generar_reporte_presentes_no_presentes(df_sire_sunat, df_sire_bn, nombre_sire_sunat, nombre_sire_bn, ruta_salida):
    """
    Genera un reporte Excel que muestra, por tipo de documento:
    - Cuántos registros del SIRE_BN (archivo 2) están presentes en SIRE_SUNAT (archivo 1)
    - Cuántos NO están presentes en SIRE_SUNAT
    - Porcentajes de presentes y no presentes (suma = 100%)
    - Nombre del tipo de documento
    - Barra de datos en todas las columnas de porcentaje (mantenido con openpyxl)
    - Detalle de TODAS las filas para cada tipo (incluyendo duplicados)
    - Descripción explicativa debajo de la tabla
    - Gráfica de barras agrupadas como imagen (con matplotlib)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.drawing.image import Image
    import matplotlib.pyplot as plt
    import io
    import pandas as pd

    required_cols = ['ID_CONCILIACION', 'TipoDoc_Norm']
    for col in required_cols:
        if col not in df_sire_sunat.columns or col not in df_sire_bn.columns:
            raise ValueError(f"Ambos DataFrames deben tener la columna '{col}'")

    # Diccionario de nombres de tipos de documento
    TIPOS_DOC_NOMBRES = {
        '00': 'Sin tipo / Otros',
        '01': 'Factura',
        '02': 'Nota de débito',
        '03': 'Boleta',
        '04': 'Nota de crédito',
        '05': 'Pasajes aéreos',
        '06': 'Comprobante de retención',
        '07': 'Nota de crédito',
        '08': 'Nota de débito',
        '09': 'Pagos exterior',
        '12': 'Ticket máquina registradora',
        '14': 'Recibo de servicio público',
        '16': 'Boleto de viaje / Transporte de carga',
        '30': 'Documentos autorizados',
        '42': 'Otros',
        '46': 'No domiciliado',
        '53': 'Declaración de mensajería o courier',
        '87': 'Nota de crédito especial',
        'DESCONOCIDO': 'Tipo desconocido'
    }

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- HOJA 1: Resumen ----------
    ws_resumen = wb.create_sheet("Resumen")

    # TÍTULO
    ws_resumen['A1'] = 'RESUMEN: REGISTROS DEL SIRE_BN'
    ws_resumen.merge_cells('A1:H1')
    ws_resumen['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_resumen['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # ENCABEZADOS CON NUEVAS COLUMNAS
    headers = [
        'Tipo de Documento',
        'Nombre de Tipo de Doc',
        'Presentes en SIRE_SUNAT',
        'Coincidencia(%)',
        'No presentes en SIRE_SUNAT',
        'Coincidencia(%)',
        'Total en SIRE_BN',
        'Coincidencia(%)'
    ]
    ws_resumen.append(headers)

    df_sire_bn['TipoDoc_Norm'] = df_sire_bn['TipoDoc_Norm'].fillna('DESCONOCIDO')
    df_sire_bn['TipoDoc_Norm'] = df_sire_bn['TipoDoc_Norm'].replace('', 'DESCONOCIDO')
    
    tipos_sire_bn = df_sire_bn['TipoDoc_Norm'].unique()
    tipos_numericos = sorted([t for t in tipos_sire_bn if t != 'DESCONOCIDO' and t.isdigit()], key=lambda x: int(x))
    tipos_ordenados = tipos_numericos + (['DESCONOCIDO'] if 'DESCONOCIDO' in tipos_sire_bn else [])

    ids_sire_sunat = set(df_sire_sunat['ID_CONCILIACION'].dropna())

    fila_actual = 3
    total_presentes = 0
    total_no_presentes = 0
    total_registros_sire_bn = 0
    
    # Guardar datos para la gráfica
    datos_grafica = []

    for tipo in tipos_ordenados:
        df_sire_bn_tipo = df_sire_bn[df_sire_bn['TipoDoc_Norm'] == tipo]
        total_registros = len(df_sire_bn_tipo)
        
        mask_presente = df_sire_bn_tipo['ID_CONCILIACION'].isin(ids_sire_sunat)
        presentes_count = mask_presente.sum()
        no_presentes_count = total_registros - presentes_count
        
        # Calcular porcentajes (como decimales)
        pct_presentes = (presentes_count / total_registros) if total_registros > 0 else 0
        pct_no_presentes = (no_presentes_count / total_registros) if total_registros > 0 else 0
        pct_total = 1.0  # Siempre 100%

        total_presentes += presentes_count
        total_no_presentes += no_presentes_count
        total_registros_sire_bn += total_registros

        nombre_tipo = TIPOS_DOC_NOMBRES.get(tipo, tipo)
        
        # Guardar para gráfica
        datos_grafica.append({
            'Tipo': tipo,
            'Nombre': nombre_tipo,
            'Presentes': int(presentes_count),
            'No_Presentes': int(no_presentes_count),
            'Total': int(total_registros)
        })

        ws_resumen.cell(row=fila_actual, column=1, value=tipo)
        ws_resumen.cell(row=fila_actual, column=2, value=nombre_tipo)
        ws_resumen.cell(row=fila_actual, column=3, value=int(presentes_count))
        ws_resumen.cell(row=fila_actual, column=4, value=pct_presentes)
        ws_resumen.cell(row=fila_actual, column=5, value=int(no_presentes_count))
        ws_resumen.cell(row=fila_actual, column=6, value=pct_no_presentes)
        ws_resumen.cell(row=fila_actual, column=7, value=int(total_registros))
        ws_resumen.cell(row=fila_actual, column=8, value=pct_total)
        fila_actual += 1

    # Fila de TOTAL
    pct_total_presentes = (total_presentes / total_registros_sire_bn) if total_registros_sire_bn > 0 else 0
    pct_total_no_presentes = (total_no_presentes / total_registros_sire_bn) if total_registros_sire_bn > 0 else 0
    
    ws_resumen.cell(row=fila_actual, column=1, value='TOTAL')
    ws_resumen.cell(row=fila_actual, column=2, value='')
    ws_resumen.cell(row=fila_actual, column=3, value=int(total_presentes))
    ws_resumen.cell(row=fila_actual, column=4, value=pct_total_presentes)
    ws_resumen.cell(row=fila_actual, column=5, value=int(total_no_presentes))
    ws_resumen.cell(row=fila_actual, column=6, value=pct_total_no_presentes)
    ws_resumen.cell(row=fila_actual, column=7, value=int(total_registros_sire_bn))
    ws_resumen.cell(row=fila_actual, column=8, value=1.0)

    _aplicar_formato_encabezado(ws_resumen, row=2)
    
    # Negrita para la fila de total
    for col_idx in range(1, 9):
        ws_resumen.cell(row=fila_actual, column=col_idx).font = Font(bold=True)

    # Formato de porcentaje para columnas D, F, H (índices 4, 6, 8)
    for row in range(3, fila_actual + 1):
        for col_idx in [4, 6, 8]:
            cell = ws_resumen.cell(row=row, column=col_idx)
            cell.number_format = '0.00%'

    # ---- BARRAS DE DATOS para columnas de porcentaje (D, F, H) ----
    data_bar_rule = DataBarRule(
        start_type='min',
        end_type='max',
        color='2563EB',
        showValue=True,
        minLength=None,
        maxLength=None
    )
    if fila_actual > 3:
        # Columna D (Presentes %)
        ws_resumen.conditional_formatting.add(
            f'D3:D{fila_actual-1}',
            data_bar_rule
        )
        # Columna F (No presentes %)
        ws_resumen.conditional_formatting.add(
            f'F3:F{fila_actual-1}',
            data_bar_rule
        )
        # Columna H (Total %)
        ws_resumen.conditional_formatting.add(
            f'H3:H{fila_actual-1}',
            data_bar_rule
        )

    # Ajustar ancho de columnas
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 30
    ws_resumen.column_dimensions['C'].width = 22
    ws_resumen.column_dimensions['D'].width = 18
    ws_resumen.column_dimensions['E'].width = 25
    ws_resumen.column_dimensions['F'].width = 18
    ws_resumen.column_dimensions['G'].width = 18
    ws_resumen.column_dimensions['H'].width = 18

    # ---- DESCRIPCIÓN EXPLICATIVA (3-4 líneas debajo de la tabla) ----
    fila_descripcion = fila_actual + 3

    descripcion = [
        "DESCRIPCIÓN DEL REPORTE:",
        "• Coincidencia (%): Porcentaje de registros que están en ambos sistemas.",
        "• No coincidencia (%): Porcentaje de registros que faltan en uno de los sistemas.",
    ]
    
    for i, linea in enumerate(descripcion):
        ws_resumen.cell(row=fila_descripcion + i, column=1, value=linea)
        if i == 0:
            ws_resumen.cell(row=fila_descripcion + i, column=1).font = Font(bold=True, size=12)
        ws_resumen.cell(row=fila_descripcion + i, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for i in range(len(descripcion)):
        ws_resumen.row_dimensions[fila_descripcion + i].height = 22

    # ---- GENERAR GRÁFICA CON MATPLOTLIB ----
    if datos_grafica:
        df_grafica = pd.DataFrame(datos_grafica)
        
        # Crear figura y ejes
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Configurar posición de barras
        x = range(len(df_grafica))
        width = 0.25
        
        # Barras
        ax.bar([i - width for i in x], df_grafica['Presentes'], width, label='Presentes en SIRE_SUNAT', color='#2563EB')
        ax.bar(x, df_grafica['No_Presentes'], width, label='No presentes en SIRE_SUNAT', color='#EF4444')
        ax.bar([i + width for i in x], df_grafica['Total'], width, label='Total en SIRE_BN', color='#10B981')
        
        # Configurar ejes
        ax.set_xlabel('Tipo de Documento')
        ax.set_ylabel('Cantidad de Registros')
        ax.set_title('Conciliación por Tipo de Documento')
        ax.set_xticks(x)
        ax.set_xticklabels(df_grafica['Tipo'], rotation=45, ha='right')
        ax.legend()
        
        # Ajustar layout
        plt.tight_layout()
        
        # Guardar gráfica en memoria como imagen
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        # Insertar imagen en Excel
        img = Image(img_buffer)
        # Ajustar tamaño de la imagen
        img.width = 700
        img.height = 400
        
        # Calcular posición (debajo de la descripción)
        fila_grafica = fila_descripcion + len(descripcion) + 2
        # Convertir fila a coordenada de celda (columna A)
        from openpyxl.utils import get_column_letter
        posicion = f'A{fila_grafica}'
        ws_resumen.add_image(img, posicion)

    # ---- GUARDAR ARCHIVO CON OPENPYXL (MANTIENE FORMATO) ----
    wb.save(ruta_salida)
    
    # ---------- HOJAS POR TIPO DE DOCUMENTO ----------
    # (Se mantiene igual, se agregan después de guardar el archivo principal)
    from openpyxl import load_workbook
    
    wb = load_workbook(ruta_salida)
    
    for tipo in tipos_ordenados:
        sheet_name = f"Tipo {tipo}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws_tipo = wb.create_sheet(sheet_name)

        df_sire_bn_tipo = df_sire_bn[df_sire_bn['TipoDoc_Norm'] == tipo]
        mask_presente = df_sire_bn_tipo['ID_CONCILIACION'].isin(ids_sire_sunat)

        fila_inicio = 1
        ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
        celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'PRESENTES EN SIRE_SUNAT (Total: {mask_presente.sum()})')
        celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        celda_titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        fila_inicio += 1

        df_presentes = df_sire_bn_tipo[mask_presente].copy()
        if not df_presentes.empty:
            cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                         [c for c in df_presentes.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
            df_presentes = df_presentes[cols_orden]
            _escribir_dataframe_con_formato(ws_tipo, df_presentes, start_row=fila_inicio)
            fila_inicio += len(df_presentes) + 2
        else:
            ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros presentes en SIRE_SUNAT para este tipo.')
            fila_inicio += 2

        ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
        celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'NO PRESENTES EN SIRE_SUNAT (Total: {(~mask_presente).sum()})')
        celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        celda_titulo.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        fila_inicio += 1

        # --- Sección: No presentes en SIRE_SUNAT (con separación por mes) ---
        df_no_presentes = df_sire_bn_tipo[~mask_presente].copy()

        if not df_no_presentes.empty:
            fechas = df_no_presentes.iloc[:, 4] if df_no_presentes.shape[1] > 4 else pd.Series([''] * len(df_no_presentes))
            meses = fechas.apply(extraer_mes_fecha)
            meses_validos = [m for m in meses if m is not None]
            mes_actual = max(meses_validos) if meses_validos else None

            if mes_actual:
                mask_mes_actual = meses == mes_actual
                df_mes_actual = df_no_presentes[mask_mes_actual].copy()
                df_meses_anteriores = df_no_presentes[~mask_mes_actual].copy()

                if not df_mes_actual.empty:
                    df_mes_actual = df_mes_actual.sort_values(by=df_mes_actual.columns[4], ascending=False)
                    ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                    celda_mes = ws_tipo.cell(row=fila_inicio, column=1, value=f'MES ACTUAL ({mes_actual[:4]}-{mes_actual[4:6]}) - Total: {len(df_mes_actual)}')
                    celda_mes.font = Font(bold=True, size=12, color="FFFFFF")
                    celda_mes.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    celda_mes.alignment = Alignment(horizontal='center', vertical='center')
                    fila_inicio += 1
                    
                    cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                                 [c for c in df_mes_actual.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                    df_mes_actual_export = df_mes_actual[cols_orden]
                    fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_mes_actual_export, fila_inicio, None)
                else:
                    ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros del mes actual.')
                    fila_inicio += 2

                if not df_meses_anteriores.empty:
                    df_meses_anteriores = df_meses_anteriores.sort_values(by=df_meses_anteriores.columns[4], ascending=False)
                    ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                    celda_meses = ws_tipo.cell(row=fila_inicio, column=1, value=f'MESES ANTERIORES - Total: {len(df_meses_anteriores)}')
                    celda_meses.font = Font(bold=True, size=12, color="FFFFFF")
                    celda_meses.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                    celda_meses.alignment = Alignment(horizontal='center', vertical='center')
                    fila_inicio += 1
                    
                    cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                                 [c for c in df_meses_anteriores.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                    df_meses_anteriores_export = df_meses_anteriores[cols_orden]
                    fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_meses_anteriores_export, fila_inicio, None)
                else:
                    ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros de meses anteriores.')
                    fila_inicio += 2
            else:
                ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'NO PRESENTES EN SIRE_SUNAT (Total: {len(df_no_presentes)})')
                celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
                celda_titulo.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
                fila_inicio += 1

                cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                             [c for c in df_no_presentes.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                df_no_presentes_export = df_no_presentes[cols_orden]
                fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_no_presentes_export, fila_inicio, None)
        else:
            ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros no presentes en SIRE_SUNAT para este tipo.')
            fila_inicio += 2

    wb.save(ruta_salida)
    return ruta_salida

# ------------------------------------------------------------
# 7. REPORTE: REGISTROS DE SIRE_SUNAT (DESDE SIRE_SUNAT HACIA SIRE_BN)
# ------------------------------------------------------------
def generar_reporte_presentes_no_presentes_sire_sunat(df_sire_sunat, df_sire_bn, nombre_sire_sunat, nombre_sire_bn, ruta_salida):
    """
    Genera un reporte Excel que muestra, por tipo de documento:
    - Cuántos registros del SIRE_SUNAT (archivo 1) están presentes en SIRE_BN (archivo 2)
    - Cuántos NO están presentes en SIRE_BN
    - Porcentajes de presentes y no presentes (suma = 100%)
    - Nombre del tipo de documento
    - Barra de datos en todas las columnas de porcentaje (mantenido con openpyxl)
    - Detalle de TODAS las filas para cada tipo (incluyendo duplicados)
    - Descripción explicativa debajo de la tabla
    - Gráfica de barras agrupadas como imagen (con matplotlib)
    
    Título: RESUMEN: REGISTROS DE SIRE_SUNAT
    Columnas: Tipo de Documento | Nombre de Tipo de Doc | Presentes en SIRE_BN | Coincidencia(%) | No presentes en SIRE_BN | Coincidencia(%) | Total en SIRE_SUNAT | Coincidencia(%)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.drawing.image import Image
    import matplotlib.pyplot as plt
    import io
    import pandas as pd

    required_cols = ['ID_CONCILIACION', 'TipoDoc_Norm']
    for col in required_cols:
        if col not in df_sire_sunat.columns or col not in df_sire_bn.columns:
            raise ValueError(f"Ambos DataFrames deben tener la columna '{col}'")

    # Diccionario de nombres de tipos de documento
    TIPOS_DOC_NOMBRES = {
        '00': 'Sin tipo / Otros',
        '01': 'Factura',
        '02': 'Nota de débito',
        '03': 'Boleta',
        '04': 'Nota de crédito',
        '05': 'Pasajes aéreos',
        '06': 'Comprobante de retención',
        '07': 'Nota de crédito',
        '08': 'Nota de débito',
        '09': 'Pagos exterior',
        '12': 'Ticket máquina registradora',
        '14': 'Recibo de servicio público',
        '16': 'Boleto de viaje / Transporte de carga',
        '30': 'Documentos autorizados',
        '42': 'Otros',
        '46': 'No domiciliado',
        '53': 'Declaración de mensajería o courier',
        '87': 'Nota de crédito especial',
        'DESCONOCIDO': 'Tipo desconocido'
    }

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- HOJA 1: Resumen ----------
    ws_resumen = wb.create_sheet("Resumen")

    # TÍTULO
    ws_resumen['A1'] = 'RESUMEN: REGISTROS DE SIRE_SUNAT'
    ws_resumen.merge_cells('A1:H1')
    ws_resumen['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_resumen['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # ENCABEZADOS CON NUEVAS COLUMNAS
    headers = [
        'Tipo de Documento',
        'Nombre de Tipo de Doc',
        'Presentes en SIRE_BN',
        'Coincidencia(%)',
        'No presentes en SIRE_BN',
        'Coincidencia(%)',
        'Total en SIRE_SUNAT',
        'Coincidencia(%)'
    ]
    ws_resumen.append(headers)

    df_sire_sunat['TipoDoc_Norm'] = df_sire_sunat['TipoDoc_Norm'].fillna('DESCONOCIDO')
    df_sire_sunat['TipoDoc_Norm'] = df_sire_sunat['TipoDoc_Norm'].replace('', 'DESCONOCIDO')
    
    tipos_sire_sunat = df_sire_sunat['TipoDoc_Norm'].unique()
    tipos_numericos = sorted([t for t in tipos_sire_sunat if t != 'DESCONOCIDO' and t.isdigit()], key=lambda x: int(x))
    tipos_ordenados = tipos_numericos + (['DESCONOCIDO'] if 'DESCONOCIDO' in tipos_sire_sunat else [])

    ids_sire_bn = set(df_sire_bn['ID_CONCILIACION'].dropna())

    fila_actual = 3
    total_presentes = 0
    total_no_presentes = 0
    total_registros_sire_sunat = 0
    
    # Guardar datos para la gráfica
    datos_grafica = []

    for tipo in tipos_ordenados:
        df_sire_sunat_tipo = df_sire_sunat[df_sire_sunat['TipoDoc_Norm'] == tipo]
        total_registros = len(df_sire_sunat_tipo)
        
        mask_presente = df_sire_sunat_tipo['ID_CONCILIACION'].isin(ids_sire_bn)
        presentes_count = mask_presente.sum()
        no_presentes_count = total_registros - presentes_count
        
        # Calcular porcentajes (como decimales)
        pct_presentes = (presentes_count / total_registros) if total_registros > 0 else 0
        pct_no_presentes = (no_presentes_count / total_registros) if total_registros > 0 else 0
        pct_total = 1.0  # Siempre 100%

        total_presentes += presentes_count
        total_no_presentes += no_presentes_count
        total_registros_sire_sunat += total_registros

        nombre_tipo = TIPOS_DOC_NOMBRES.get(tipo, tipo)
        
        # Guardar para gráfica
        datos_grafica.append({
            'Tipo': tipo,
            'Nombre': nombre_tipo,
            'Presentes': int(presentes_count),
            'No_Presentes': int(no_presentes_count),
            'Total': int(total_registros)
        })

        ws_resumen.cell(row=fila_actual, column=1, value=tipo)
        ws_resumen.cell(row=fila_actual, column=2, value=nombre_tipo)
        ws_resumen.cell(row=fila_actual, column=3, value=int(presentes_count))
        ws_resumen.cell(row=fila_actual, column=4, value=pct_presentes)
        ws_resumen.cell(row=fila_actual, column=5, value=int(no_presentes_count))
        ws_resumen.cell(row=fila_actual, column=6, value=pct_no_presentes)
        ws_resumen.cell(row=fila_actual, column=7, value=int(total_registros))
        ws_resumen.cell(row=fila_actual, column=8, value=pct_total)
        fila_actual += 1

    # Fila de TOTAL
    pct_total_presentes = (total_presentes / total_registros_sire_sunat) if total_registros_sire_sunat > 0 else 0
    pct_total_no_presentes = (total_no_presentes / total_registros_sire_sunat) if total_registros_sire_sunat > 0 else 0
    
    ws_resumen.cell(row=fila_actual, column=1, value='TOTAL')
    ws_resumen.cell(row=fila_actual, column=2, value='')
    ws_resumen.cell(row=fila_actual, column=3, value=int(total_presentes))
    ws_resumen.cell(row=fila_actual, column=4, value=pct_total_presentes)
    ws_resumen.cell(row=fila_actual, column=5, value=int(total_no_presentes))
    ws_resumen.cell(row=fila_actual, column=6, value=pct_total_no_presentes)
    ws_resumen.cell(row=fila_actual, column=7, value=int(total_registros_sire_sunat))
    ws_resumen.cell(row=fila_actual, column=8, value=1.0)

    _aplicar_formato_encabezado(ws_resumen, row=2)
    
    # Negrita para la fila de total
    for col_idx in range(1, 9):
        ws_resumen.cell(row=fila_actual, column=col_idx).font = Font(bold=True)

    # Formato de porcentaje para columnas D, F, H (índices 4, 6, 8)
    for row in range(3, fila_actual + 1):
        for col_idx in [4, 6, 8]:
            cell = ws_resumen.cell(row=row, column=col_idx)
            cell.number_format = '0.00%'

    # ---- BARRAS DE DATOS para columnas de porcentaje (D, F, H) ----
    data_bar_rule = DataBarRule(
        start_type='min',
        end_type='max',
        color='2563EB',
        showValue=True,
        minLength=None,
        maxLength=None
    )
    if fila_actual > 3:
        # Columna D (Presentes %)
        ws_resumen.conditional_formatting.add(
            f'D3:D{fila_actual-1}',
            data_bar_rule
        )
        # Columna F (No presentes %)
        ws_resumen.conditional_formatting.add(
            f'F3:F{fila_actual-1}',
            data_bar_rule
        )
        # Columna H (Total %)
        ws_resumen.conditional_formatting.add(
            f'H3:H{fila_actual-1}',
            data_bar_rule
        )

    # Ajustar ancho de columnas
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 30
    ws_resumen.column_dimensions['C'].width = 22
    ws_resumen.column_dimensions['D'].width = 18
    ws_resumen.column_dimensions['E'].width = 25
    ws_resumen.column_dimensions['F'].width = 18
    ws_resumen.column_dimensions['G'].width = 22
    ws_resumen.column_dimensions['H'].width = 18

    # ---- DESCRIPCIÓN EXPLICATIVA (3-4 líneas debajo de la tabla) ----
    fila_descripcion = fila_actual + 3

    descripcion = [
        "DESCRIPCIÓN DEL REPORTE:",
        "Este reporte compara los comprobantes registrados en SIRE_SUNAT y SIRE_BN.",
        "• Coincidencia (%): Porcentaje de registros que están en ambos sistemas.",
        "• No coincidencia (%): Porcentaje de registros que faltan en uno de los sistemas.",
    ]
    
    for i, linea in enumerate(descripcion):
        ws_resumen.cell(row=fila_descripcion + i, column=1, value=linea)
        if i == 0:
            ws_resumen.cell(row=fila_descripcion + i, column=1).font = Font(bold=True, size=12)
        ws_resumen.cell(row=fila_descripcion + i, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for i in range(len(descripcion)):
        ws_resumen.row_dimensions[fila_descripcion + i].height = 22

    # ---- GENERAR GRÁFICA CON MATPLOTLIB ----
    if datos_grafica:
        df_grafica = pd.DataFrame(datos_grafica)
        
        # Crear figura y ejes
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Configurar posición de barras
        x = range(len(df_grafica))
        width = 0.25
        
        # Barras
        ax.bar([i - width for i in x], df_grafica['Presentes'], width, label='Presentes en SIRE_BN', color='#2563EB')
        ax.bar(x, df_grafica['No_Presentes'], width, label='No presentes en SIRE_BN', color='#EF4444')
        ax.bar([i + width for i in x], df_grafica['Total'], width, label='Total en SIRE_SUNAT', color='#10B981')
        
        # Configurar ejes
        ax.set_xlabel('Tipo de Documento')
        ax.set_ylabel('Cantidad de Registros')
        ax.set_title('Conciliación por Tipo de Documento')
        ax.set_xticks(x)
        ax.set_xticklabels(df_grafica['Tipo'], rotation=45, ha='right')
        ax.legend()
        
        # Ajustar layout
        plt.tight_layout()
        
        # Guardar gráfica en memoria como imagen
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        # Insertar imagen en Excel
        img = Image(img_buffer)
        # Ajustar tamaño de la imagen
        img.width = 700
        img.height = 400
        
        # Calcular posición (debajo de la descripción)
        fila_grafica = fila_descripcion + len(descripcion) + 2
        # Convertir fila a coordenada de celda (columna A)
        from openpyxl.utils import get_column_letter
        posicion = f'A{fila_grafica}'
        ws_resumen.add_image(img, posicion)

    # ---- GUARDAR ARCHIVO CON OPENPYXL (MANTIENE FORMATO) ----
    wb.save(ruta_salida)
    
    # ---------- HOJAS POR TIPO DE DOCUMENTO ----------
    # (Se mantiene igual, se agregan después de guardar el archivo principal)
    from openpyxl import load_workbook
    
    wb = load_workbook(ruta_salida)
    
    for tipo in tipos_ordenados:
        sheet_name = f"Tipo {tipo}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws_tipo = wb.create_sheet(sheet_name)

        df_sire_sunat_tipo = df_sire_sunat[df_sire_sunat['TipoDoc_Norm'] == tipo]
        mask_presente = df_sire_sunat_tipo['ID_CONCILIACION'].isin(ids_sire_bn)

        fila_inicio = 1
        ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
        celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'PRESENTES EN SIRE_BN (Total: {mask_presente.sum()})')
        celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        celda_titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        fila_inicio += 1

        df_presentes = df_sire_sunat_tipo[mask_presente].copy()
        if not df_presentes.empty:
            cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                         [c for c in df_presentes.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
            df_presentes = df_presentes[cols_orden]
            _escribir_dataframe_con_formato(ws_tipo, df_presentes, start_row=fila_inicio)
            fila_inicio += len(df_presentes) + 2
        else:
            ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros presentes en SIRE_BN para este tipo.')
            fila_inicio += 2

        ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
        celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'NO PRESENTES EN SIRE_BN (Total: {(~mask_presente).sum()})')
        celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        celda_titulo.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        fila_inicio += 1

        # --- Sección: No presentes en SIRE_BN (con separación por mes) ---
        df_no_presentes = df_sire_sunat_tipo[~mask_presente].copy()
        
        if not df_no_presentes.empty:
            fechas = df_no_presentes.iloc[:, 4] if df_no_presentes.shape[1] > 4 else pd.Series([''] * len(df_no_presentes))
            meses = fechas.apply(extraer_mes_fecha)
            meses_validos = [m for m in meses if m is not None]
            mes_actual = max(meses_validos) if meses_validos else None
            
            if mes_actual:
                mask_mes_actual = meses == mes_actual
                df_mes_actual = df_no_presentes[mask_mes_actual].copy()
                df_meses_anteriores = df_no_presentes[~mask_mes_actual].copy()
                
                if not df_mes_actual.empty:
                    df_mes_actual = df_mes_actual.sort_values(by=df_mes_actual.columns[4], ascending=False)
                    ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                    celda_mes = ws_tipo.cell(row=fila_inicio, column=1, value=f'MES ACTUAL ({mes_actual[:4]}-{mes_actual[4:6]}) - Total: {len(df_mes_actual)}')
                    celda_mes.font = Font(bold=True, size=12, color="FFFFFF")
                    celda_mes.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    celda_mes.alignment = Alignment(horizontal='center', vertical='center')
                    fila_inicio += 1
                    
                    cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                                 [c for c in df_mes_actual.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                    df_mes_actual_export = df_mes_actual[cols_orden]
                    fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_mes_actual_export, fila_inicio, None)
                else:
                    ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros del mes actual.')
                    fila_inicio += 2
                
                if not df_meses_anteriores.empty:
                    df_meses_anteriores = df_meses_anteriores.sort_values(by=df_meses_anteriores.columns[4], ascending=False)
                    ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                    celda_meses = ws_tipo.cell(row=fila_inicio, column=1, value=f'MESES ANTERIORES - Total: {len(df_meses_anteriores)}')
                    celda_meses.font = Font(bold=True, size=12, color="FFFFFF")
                    celda_meses.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                    celda_meses.alignment = Alignment(horizontal='center', vertical='center')
                    fila_inicio += 1
                    
                    cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                                 [c for c in df_meses_anteriores.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                    df_meses_anteriores_export = df_meses_anteriores[cols_orden]
                    fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_meses_anteriores_export, fila_inicio, None)
                else:
                    ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros de meses anteriores.')
                    fila_inicio += 2
            else:
                ws_tipo.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
                celda_titulo = ws_tipo.cell(row=fila_inicio, column=1, value=f'NO PRESENTES EN SIRE_BN (Total: {len(df_no_presentes)})')
                celda_titulo.font = Font(bold=True, size=12, color="FFFFFF")
                celda_titulo.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
                fila_inicio += 1
                
                cols_orden = ['ID_CONCILIACION', 'Hoja_Origen'] + \
                             [c for c in df_no_presentes.columns if c not in ['ID_CONCILIACION', 'Hoja_Origen']]
                df_no_presentes_export = df_no_presentes[cols_orden]
                fila_inicio = _escribir_dataframe_con_formato_con_titulo(ws_tipo, df_no_presentes_export, fila_inicio, None)
        else:
            ws_tipo.cell(row=fila_inicio, column=1, value='No hay registros no presentes en SIRE_BN para este tipo.')
            fila_inicio += 2

    wb.save(ruta_salida)
    return ruta_salida
