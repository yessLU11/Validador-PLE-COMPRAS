# ============================================================================
# duplicate_report_generator_internal.py
# ============================================================================
# Descripción: Genera reporte Excel con duplicados encontrados dentro de un mismo archivo.
# Crea archivo con dos hojas:
#   1. Duplicados_Detalle: Todas las columnas + hoja_origen + fila_original
#   2. Auditoria_Resumen: Estadísticas de duplicados
# Se relaciona con excel_reader.py para obtener los datos leídos desde Excel,
# con database.py para cargar datos históricos y con report_generator.py para generar el reporte final.
# ============================================================================

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def generar_reporte_duplicados_interno(
    df_original: pd.DataFrame,
    df_duplicados: pd.DataFrame,
    auditoria: dict,
    nombre_archivo_salida: str = "duplicados.xlsx",
    nombre_archivo_original: str = "archivo.xlsx"
) -> str:
    """
    Genera un archivo Excel con los duplicados encontrados.
    
    Parámetros:
    - df_original: DataFrame con TODOS los datos originales del Excel
    - df_duplicados: DataFrame con solo los registros duplicados (con metadata)
    - auditoria: Dict con estadísticas
    - nombre_archivo_salida: Nombre del archivo a generar
    - nombre_archivo_original: Nombre del archivo original (para contexto)
    
    Retorna: Ruta del archivo generado
    """
    
    # Preparar datos para la hoja de detalle
    df_detalle = df_duplicados.copy()
    
    # Reordenar columnas: metadata primero, luego datos originales
    cols_metadata = ["Hoja_origen", "Fila Original"]
    
    # Columnas de datos originales (excluir las normalizadas y metadata)
    cols_datos = [col for col in df_detalle.columns 
                  if not (isinstance(col, str) and col.startswith("_")) and col not in cols_metadata]
    
    # Orden final: metadata + datos, esto me indica que las columnas de datos originales se mantengan en el mismo orden que en el DataFrame original, pero siempre después de la metadata.
    orden_cols = cols_metadata + cols_datos
    orden_cols = [col for col in orden_cols if col in df_detalle.columns]
    
    df_detalle = df_detalle[orden_cols] # Esto asegura que las columnas de datos originales se mantengan en el mismo orden que en el DataFrame original, pero siempre después de la metadata.
    
    # Renombrar columnas para el reporte
    # Para colocar los nombres de las columnas despues siguientes a hoja origen y fila original ¿Cómo lo hago? 
    # ¿cómo hago para tener el nombre de las columnas en el sgte orden: Hoja Origen	Fila Original	Descripción	Periodo (AAAAMM00)	Código Único de la Operación (CUO)	Número correlativo del asiento contable identificado en el campo 2, cuando se utilice el CUO.	Fecha de emisión del comprobante de pago o documento	Fecha de Vencimiento o Fecha de Pago	Tipo de Comprobante de Pago o Documento	Serie del comprobante de pago o documento.	Año de emisión de la DUA o DSI	Número del comprobante de pago o documento o número de orden del formulario.	En caso de optar por anotar el importe total de las operaciones diarias...en forma consolidada, registrar el número final	Tipo de Documento de Identidad del proveedor	Número de RUC del proveedor o número de documento de Identidad, según corresponda.	Apellidos y nombres, denominación o razón social  del proveedor. En caso de personas naturales se debe consignar los datos en el siguiente orden: apellido paterno, apellido materno y nombre completo.	Base imponible de las adquisiciones gravadas ... destinadas exclusivamente a operaciones gravadas y/o de exportación	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Base imponible de las adquisiciones gravadas ... destinadas a operaciones gravadas y/o de exportación y a operaciones no gravadas	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Base imponible de las adquisiciones gravadas que no dan derecho a crédito fiscal …	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Valor de las adquisiciones no gravadas	Monto del Impuesto Selectivo al Consumo en los casos en que el sujeto pueda utilizarlo como deducción.	Impuesto al consumo de las bolsas de plastico	Otros conceptos, tributos y cargos que no formen parte de la base imponible.	Importe total de las adquisiciones registradas según comprobante de pago.	Código de la Moneda (Tabla 4)	Tipo de cambio	Fecha de emisión del comprobante de pago que se modifica	Tipo de comprobante de pago que se modifica	Número de serie del comprobante de pago que se modifica	Código dependencia aduanera	Número del comprobante de pago que se modifica	Fecha de emisión de la Constancia de Depósito de Detracción 	Número de la Constancia de Depósito de Detracción	Marca del comprobante de pago sujeto a retención	Clasificación de los bienes y servicios adquiridos (Tabla 30)	Identificación del Contrato o del proyecto en el caso de los Operadores de las sociedades irregulares	Error tipo 1: inconsistencia en el tipo de cambio	Error tipo 2: inconsistencia por proveedores no habidos	Error tipo 3: inconsistencia por proveedores que renunciaron a la exoneración del Apéndice I del IGV	Error tipo 4: inconsistencia por DNIs que fueron utilizados en las Liquidaciones de Compra	Indicador de Comprobantes de pago cancelados con medios de pago	Estado que identifica la oportunidad de la anotación o indicación si ésta corresponde a un ajuste.? 
    # Para esto, lo que hago es crear un diccionario de mapeo de nombres de columnas originales a los nombres que quiero mostrar en el reporte, y luego uso ese diccionario para renombrar las columnas del DataFrame antes de exportarlo a Excel.
    # Paso 1: Crear diccionario de mapeo
    # Paso 2: Renombrar columnas usando el diccionario de mapeo, esto me permite tener un control total sobre cómo se muestran los nombres de las columnas en el reporte final, asegurando que sean claros y profesionales para el usuario final.
    # Desde la columna 3 deben ir los nombres de la fila 2 del archivo original  que son : Descripción	Periodo (AAAAMM00)	Código Único de la Operación (CUO)	Número correlativo del asiento contable identificado en el campo 2, cuando se utilice el CUO.	Fecha de emisión del comprobante de pago o documento	Fecha de Vencimiento o Fecha de Pago	Tipo de Comprobante de Pago o Documento	Serie del comprobante de pago o documento.	Año de emisión de la DUA o DSI	Número del comprobante de pago o documento o número de orden del formulario.	En caso de optar por anotar el importe total de las operaciones diarias...en forma consolidada, registrar el número final	Tipo de Documento de Identidad del proveedor	Número de RUC del proveedor o número de documento de Identidad, según corresponda.	Apellidos y nombres, denominación o razón social  del proveedor. En caso de personas naturales se debe consignar los datos en el siguiente orden: apellido paterno, apellido materno y nombre completo.	Base imponible de las adquisiciones gravadas ... destinadas exclusivamente a operaciones gravadas y/o de exportación	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Base imponible de las adquisiciones gravadas ... destinadas a operaciones gravadas y/o de exportación y a operaciones no gravadas	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Base imponible de las adquisiciones gravadas que no dan derecho a crédito fiscal …	Monto del Impuesto General a las Ventas y/o Impuesto de Promoción Municipal	Valor de las adquisiciones no gravadas	Monto del Impuesto Selectivo al Consumo en los casos en que el sujeto pueda utilizarlo como deducción.	Impuesto al consumo de las bolsas de plastico	Otros conceptos, tributos y cargos que no formen parte de la base imponible.	Importe total de las adquisiciones registradas según comprobante de pago.	Código de la Moneda (Tabla 4)	Tipo de cambio	Fecha de emisión del comprobante de pago que se modifica	Tipo de comprobante de pago que se modifica	Número de serie del comprobante de pago que se modifica	Código dependencia aduanera	Número del comprobante de pago que se modifica	Fecha de emisión de la Constancia de Depósito de Detracción 	Número de la Constancia de Depósito de Detracción	Marca del comprobante de pago sujeto a retención	Clasificación de los bienes y servicios adquiridos (Tabla 30)	Identificación del Contrato o del proyecto en el caso de los Operadores de las sociedades irregulares	Error tipo 1: inconsistencia en el tipo de cambio	Error tipo 2: inconsistencia por proveedores no habidos	Error tipo 3: inconsistencia por proveedores que renunciaron a la exoneración del Apéndice I del IGV	Error tipo 4: inconsistencia por DNIs que fueron utilizados en las Liquidaciones de Compra	Indicador de Comprobantes de pago cancelados con medios de pago	Estado que identifica la oportunidad de la anotación o indicación si ésta corresponde a un ajuste.		
    # Para esto, lo que hago es crear un diccionario de mapeo de nombres de columnas originales a los nombres que quiero mostrar en el reporte, y luego uso ese diccionario para renombrar las columnas del DataFrame antes de exportarlo a Excel.
    # Paso 1: Crear diccionario de mapeo    
    mapeo_nombres = {
        "hoja_origen": "Hoja Origen",
        "numero_fila_original": "Fila Original",
    }

    # Renombrar solo las columnas de metadata
    rename_meta = {
        "hoja_origen": "Hoja Origen",
        "numero_fila_original": "Fila Original"
    }
    df_detalle = df_detalle.rename(columns=rename_meta, errors="ignore")
    
    # Crear archivo Excel
    with pd.ExcelWriter(nombre_archivo_salida, engine="openpyxl") as writer:
        # ===== HOJA 1: Detalle de Duplicados =====
        df_detalle.to_excel(
            writer, 
            sheet_name="Duplicados_Detalle", 
            index=False
        )
        
        # ===== HOJA 2: Auditoría/Resumen =====
        resumen_data = {
            "Métrica": [
                "Total de filas leídas",
                "Total de registros duplicados",
                "Grupos únicos de duplicados",
                "Archivo original",
                "Fecha de procesamiento"
            ],
            "Valor": [
                auditoria.get("total_filas", 0),
                auditoria.get("total_duplicados", 0),
                auditoria.get("grupos_duplicados", 0),
                nombre_archivo_original,
                pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(
            writer,
            sheet_name="Auditoria_Resumen",
            index=False
        )
        
        # ===== HOJA 3: Duplicados por Hoja =====
        if auditoria.get("duplicados_por_hoja"):
            duplicados_por_hoja = [
                {"Hoja": hoja, "Cantidad de Duplicados": cantidad}
                for hoja, cantidad in auditoria["duplicados_por_hoja"].items()
            ]
            df_por_hoja = pd.DataFrame(duplicados_por_hoja)
            df_por_hoja.to_excel(
                writer,
                sheet_name="Resumen_por_Hoja",
                index=False
            )
    
    # Aplicar estilos con openpyxl
    _aplicar_estilos(nombre_archivo_salida)
    
    return nombre_archivo_salida


def _aplicar_estilos(ruta_archivo: str):
    """Aplica estilos profesionales al archivo Excel generado."""
    
    try:
        wb = load_workbook(ruta_archivo)
        
        # Definir estilos
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")  # Azul
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        duplicado_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Rojo claro
        duplicado_font = Font(color="991B1B", bold=True)  # Rojo oscuro para texto de duplicados
        
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Aplicar estilos a "Duplicados_Detalle"
        if "Duplicados_Detalle" in wb.sheetnames:
            ws = wb["Duplicados_Detalle"]
            
            # Header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = left_alignment
                    # Resaltar filas con duplicados
                    if cell.row > 1:
                        cell.fill = duplicado_fill
                        cell.font = duplicado_font
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Max 50 para legibilidad
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Aplicar estilos a "Auditoria_Resumen"
        if "Auditoria_Resumen" in wb.sheetnames:
            ws = wb["Auditoria_Resumen"]
            
            # Header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border 
                    cell.alignment = left_alignment
            
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 50
        
        # Aplicar estilos a "Resumen_por_Hoja"
        if "Resumen_por_Hoja" in wb.sheetnames:
            ws = wb["Resumen_por_Hoja"]
            
            # Header: Esta parte es opcional, pero si existe esta hoja, le damos el mismo estilo profesional.
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos: Aquí también aplicamos bordes y alineación a los datos de esta hoja.
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_alignment
            
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
        
        wb.save(ruta_archivo)
        
    except Exception as e:
        print(f"Advertencia: No se pudieron aplicar estilos al Excel: {e}")


def generar_archivo_descarga(
    df_duplicados: pd.DataFrame,
    auditoria: dict,
    nombre_archivo_original: str
) -> str:
    """
    Wrapper que genera un archivo con nombre automático:
    duplicados_NOMBRE_ORIGINAL_SIN_EXTENSIÓN.xlsx
    """
    
    # Extraer nombre base sin extensión
    nombre_base = os.path.splitext(nombre_archivo_original)[0]
    nombre_salida = f"duplicados_{nombre_base}.xlsx"
    
    # Si ya existe, agregar timestamp
    if os.path.exists(nombre_salida):
        import time
        timestamp = int(time.time())
        nombre_base_ts = os.path.splitext(nombre_salida)[0]
        nombre_salida = f"{nombre_base_ts}.xlsx"
    
    return nombre_salida
